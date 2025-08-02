import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from .data_analyzer import get_monthly_attendance_data, get_students_with_attendance


class ExcelReportGenerator:
    """Excel形式の出席レポート生成クラス"""
    
    def __init__(self):
        self.workbook = None
        self.output_dir = Path(__file__).parent.parent / "output" / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def create_workbook(self) -> Workbook:
        """新しいワークブックを作成"""
        self.workbook = Workbook()
        # デフォルトシートを削除
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        return self.workbook
    
    def setup_worksheet_layout(self, worksheet, student_name: str, year: int, month: int):
        """ワークシートのレイアウトを設定（A4横向き）"""
        # ページ設定
        worksheet.page_setup.orientation = 'landscape'  # 横向き
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToHeight = 1
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.scale = 85  # 1ページに収まるよう85%にスケール
        
        # マージン設定（テンプレート微調整：左余白2cm）
        worksheet.page_margins = PageMargins(
            left=0.787, right=0.5, top=0.75, bottom=0.75,  # 左余白2cm = 0.787インチ
            header=0.3, footer=0.3
        )
        
        # 列幅設定（テンプレート微調整に準拠）
        column_widths = {
            'A': 15,  # 出席日
            'B': 25,  # 利用時間
            'C': 12,  # 合計
            'D': 10,  # 気分
            'E': 10,  # 睡眠
            'F': 15,  # 目的
            'G': 15,  # プランニング（微調整で15に統一）
            'H': 15,  # カウンセリング（微調整で15に統一）
            'I': 15,  # 個別対応（微調整で15に統一）
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    
    def add_title_and_header(self, worksheet, year: int, month: int, student_name: str):
        """レポートタイトル、ロゴ、生徒氏名を追加（テンプレート形式完全準拠）"""
        # タイトルを1行目中央に配置（テンプレート形式完全準拠）
        title = f"{month}月の出席レポート"
        worksheet['A1'] = title
        
        # タイトルスタイル（テンプレート完全準拠）
        title_font = Font(name='UD デジタル 教科書体 NK', size=24, bold=False)
        title_alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A1'].font = title_font
        worksheet['A1'].alignment = title_alignment
        
        # タイトル行をマージ（A1からI1まで）
        worksheet.merge_cells('A1:I1')
        
        # タイトル行の高さを調整（テンプレート準拠）
        worksheet.row_dimensions[1].height = 35.1
        
        # 生徒氏名を2行目右寄せに配置（テンプレート形式完全準拠）
        student_text = f"氏名: {student_name}"
        worksheet['A2'] = student_text
        
        # 生徒氏名スタイル（テンプレート完全準拠）
        student_font = Font(name='UD デジタル 教科書体 NK', size=20, bold=False)
        student_alignment = Alignment(horizontal='right', vertical='center')
        worksheet['A2'].font = student_font
        worksheet['A2'].alignment = student_alignment
        
        # 生徒氏名行をマージ（A2からI2まで）
        worksheet.merge_cells('A2:I2')
        
        # 生徒氏名行の高さを調整（テンプレート準拠）
        worksheet.row_dimensions[2].height = 30.0
        
        # ロゴ画像を配置
        self.add_logo(worksheet)
        
        # 3行目（空行）の高さを調整
        worksheet.row_dimensions[3].height = 10
    
    def add_logo(self, worksheet):
        """ロゴ画像を配置（縦横比固定、高さ62%）"""
        try:
            logo_path = Path(__file__).parent.parent / "assets" / "images" / "Onedrop_logo" / "Onedrop_logo_transparent.png"
            if logo_path.exists():
                img = Image(str(logo_path))
                # 元の画像サイズを取得
                original_width = img.width
                original_height = img.height
                
                # 高さを62%に設定
                new_height = int(original_height * 0.62)
                
                # 縦横比を維持して幅を計算
                aspect_ratio = original_width / original_height
                new_width = int(new_height * aspect_ratio)
                
                # 新しいサイズを適用
                img.width = new_width
                img.height = new_height
                
                # A1セルの左上に配置
                worksheet.add_image(img, 'A1')
        except Exception as e:
            print(f"ロゴ画像の配置でエラーが発生しました: {e}")
    
    def add_table_headers(self, worksheet):
        """表のヘッダーを追加"""
        headers = [
            "出席日",
            "利用時間（出席時間ー退出時間）", 
            "合計（滞在時間）",
            "気分",
            "睡眠", 
            "目的",
            "プランニング",
            "カウンセリング",
            "個別対応"
        ]
        
        # ヘッダーの設定（テンプレート完全準拠）
        header_font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行は4行目に配置（タイトル、ロゴ、空行の後）
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # ヘッダー行の高さを調整（テンプレート準拠）
        worksheet.row_dimensions[4].height = 45.0
    
    def add_attendance_data(self, worksheet, daily_records: List[Dict], start_row: int = 5):
        """出席データを表に追加"""
        if not daily_records:
            return start_row
        
        # データスタイル（文字サイズを大きく）
        data_font = Font(name='メイリオ', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = start_row
        
        for record in daily_records:
            # 出席日
            date_obj = datetime.strptime(record['date'], '%Y-%m-%d')
            date_str = date_obj.strftime('%m/%d (%a)')
            worksheet.cell(row=current_row, column=1).value = date_str
            
            # 利用時間
            time_range = f"{record['entry_time']}ー{record['exit_time']}"
            worksheet.cell(row=current_row, column=2).value = time_range
            
            # 合計（滞在時間）
            stay_time = f"{record['stay_minutes']}分"
            worksheet.cell(row=current_row, column=3).value = stay_time
            
            # 気分
            mood = record.get('mood', '')
            worksheet.cell(row=current_row, column=4).value = mood
            
            # 睡眠
            sleep = record.get('sleep_satisfaction', '')
            worksheet.cell(row=current_row, column=5).value = sleep
            
            # 目的
            purpose = record.get('purpose', '')
            worksheet.cell(row=current_row, column=6).value = purpose
            
            # プランニング、カウンセリング、個別対応は空欄（後で手動入力）
            worksheet.cell(row=current_row, column=7).value = ''
            worksheet.cell(row=current_row, column=8).value = ''
            worksheet.cell(row=current_row, column=9).value = ''
            
            # スタイル適用
            for col in range(1, 10):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
            
            # 行の高さを1ページに収まるよう調整
            worksheet.row_dimensions[current_row].height = 25
            
            current_row += 1
        
        return current_row
    
    def add_summary(self, worksheet, attendance_count: int, start_row: int):
        """利用日数の報告を追加（テンプレート形式に準拠）"""
        summary_text = f"計{attendance_count}日利用"
        worksheet.cell(row=start_row + 1, column=1).value = summary_text
        
        # サマリースタイル（文字サイズを大きく）
        summary_font = Font(name='メイリオ', size=13, bold=True)
        worksheet.cell(row=start_row + 1, column=1).font = summary_font
        
        # サマリー行の高さを1ページに収まるよう調整
        worksheet.row_dimensions[start_row + 1].height = 25
        
        return start_row + 3  # 空行を1行追加
    
    def add_comment_section(self, worksheet, start_row: int):
        """コメント欄を追加（テンプレート形式に準拠）"""
        # コメント欄タイトル（テンプレート形式に準拠）
        comment_title = "様子のコメント："
        worksheet.cell(row=start_row, column=1).value = comment_title
        
        # コメントタイトルスタイル
        comment_font = Font(name='メイリオ', size=13, bold=True)
        worksheet.cell(row=start_row, column=1).font = comment_font
        
        # タイトル行の高さを調整
        worksheet.row_dimensions[start_row].height = 25
        
        # コメント入力エリア（テンプレート形式では空行で表現）
        comment_start_row = start_row + 1
        comment_rows = 3
        
        # 3行分のセルを全て結合してテキストボックスを作成
        worksheet.merge_cells(f'A{comment_start_row}:I{comment_start_row + comment_rows - 1}')
        
        # 結合されたセルを取得
        merged_cell = worksheet.cell(row=comment_start_row, column=1)
        merged_cell.value = ''  # 空欄
        
        # コメント欄のスタイル（テンプレート形式に準拠）
        comment_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        merged_cell.border = comment_border
        merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        merged_cell.font = Font(name='メイリオ', size=11)
        
        # 各行の高さを調整（1ページに収まるよう）
        for i in range(comment_rows):
            worksheet.row_dimensions[comment_start_row + i].height = 20
        
        return comment_start_row + comment_rows
    
    def add_dropdown_validation(self, worksheet, record_count: int):
        """プランニング、カウンセリング、個別対応列にプルダウン検証を追加"""
        # プルダウンの選択肢を設定（「　」は空白、「〇」は丸）
        dropdown_options = '"　,〇"'
        
        # データ検証オブジェクトを作成
        data_validation = DataValidation(
            type="list",
            formula1=dropdown_options,
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True,
            error='無効な値です。「　」または「〇」を選択してください。',
            errorTitle='入力エラー'
        )
        
        # 対象列の定義
        # G列: プランニング、H列: カウンセリング、I列: 個別対応
        target_columns = ['G', 'H', 'I']
        
        # 各列の出席データ行（5行目から）にデータ検証を適用
        for col in target_columns:
            for row in range(5, 5 + record_count):
                cell_range = f"{col}{row}"
                data_validation.add(cell_range)
        
        # ワークシートにデータ検証を追加
        worksheet.add_data_validation(data_validation)
    
    def create_student_sheet(self, student_id: str, year: int, month: int) -> str:
        """生徒個人のシートを作成"""
        # 出席データを取得
        attendance_data = get_monthly_attendance_data(student_id, year, month)
        student_name = attendance_data["student_name"]
        daily_records = attendance_data["daily_records"]
        attendance_count = attendance_data["attendance_count"]
        
        # ワークシートを作成
        sheet_name = f"{student_name}_{month}月"
        # シート名の文字数制限と無効文字の置換
        safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_'))[:31]
        worksheet = self.workbook.create_sheet(title=safe_sheet_name)
        
        # レイアウト設定
        self.setup_worksheet_layout(worksheet, student_name, year, month)
        
        # タイトル、ロゴ、生徒氏名追加
        self.add_title_and_header(worksheet, year, month, student_name)
        
        # 表ヘッダー追加
        self.add_table_headers(worksheet)
        
        # 出席データ追加
        next_row = self.add_attendance_data(worksheet, daily_records)
        
        # プルダウン設定を追加（プランニング、カウンセリング、個別対応列）
        self.add_dropdown_validation(worksheet, len(daily_records))
        
        # サマリー追加（テンプレート形式に準拠）
        next_row = self.add_summary(worksheet, attendance_count, next_row)
        
        # コメント欄追加（テンプレート形式に準拠）
        self.add_comment_section(worksheet, next_row)
        
        return safe_sheet_name
    
    def generate_monthly_reports(self, year: int, month: int) -> str:
        """指定月の全生徒のレポートを1つのExcelファイルに生成"""
        try:
            # ワークブック作成
            self.create_workbook()
            
            # 対象月に出席記録がある生徒を取得
            students = get_students_with_attendance(year, month)
            
            if not students:
                print(f"{year}年{month}月に出席記録がある生徒はいません")
                return ""
            
            generated_sheets = []
            
            # 生徒ごとにシートを作成
            for student in students:
                try:
                    print(f"生徒 {student['name']} ({student['id']}) のシートを作成中...")
                    sheet_name = self.create_student_sheet(student["id"], year, month)
                    generated_sheets.append(sheet_name)
                    print(f"[OK] 完了: {student['name']}")
                except Exception as e:
                    print(f"[ERROR] エラー: {student['name']} - {e}")
                    continue
            
            # ファイル名生成
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H-%M-%S")
            filename = f"{month}月の出席レポート_{date_str}.{time_str}.xlsx"
            output_path = self.output_dir / filename
            
            # Excelファイル保存
            self.workbook.save(str(output_path))
            
            print(f"\nExcel レポート生成完了")
            print(f"ファイル: {output_path}")
            print(f"生成シート数: {len(generated_sheets)}件")
            print(f"対象生徒: {', '.join(generated_sheets)}")
            
            return str(output_path)
            
        except Exception as e:
            print(f"Excel レポート生成中にエラーが発生しました: {e}")
            raise
    
    def generate_single_student_report(self, student_id: str, year: int, month: int) -> str:
        """単一生徒のExcelレポートを生成"""
        try:
            # ワークブック作成
            self.create_workbook()
            
            # 生徒のシートを作成
            sheet_name = self.create_student_sheet(student_id, year, month)
            
            # 出席データを取得（生徒名取得のため）
            attendance_data = get_monthly_attendance_data(student_id, year, month)
            student_name = attendance_data["student_name"]
            
            # ファイル名生成
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_student_name = "".join(c for c in student_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            filename = f"出席レポート_{year}年{month:02d}月_{student_id}_{safe_student_name}_{timestamp}.xlsx"
            output_path = self.output_dir / filename
            
            # Excelファイル保存
            self.workbook.save(str(output_path))
            
            print(f"個人Excel レポート生成完了: {student_name}")
            print(f"ファイル: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            print(f"個人Excel レポート生成中にエラーが発生しました: {e}")
            raise


def generate_excel_reports(year: int, month: int) -> str:
    """Excel形式の月次レポートを生成（メイン関数）"""
    generator = ExcelReportGenerator()
    return generator.generate_monthly_reports(year, month)


def generate_single_excel_report(student_id: str, year: int, month: int) -> str:
    """単一生徒のExcel形式レポートを生成"""
    generator = ExcelReportGenerator()
    return generator.generate_single_student_report(student_id, year, month)


def test_excel_generation():
    """Excel生成のテスト"""
    try:
        print("Excel レポート生成テストを開始...")
        
        # テストデータ
        test_year = 2025
        test_month = 7
        
        # 全生徒のレポート生成テスト
        print("\n全生徒レポート生成テスト:")
        excel_path = generate_excel_reports(test_year, test_month)
        
        if excel_path and Path(excel_path).exists():
            file_size = Path(excel_path).stat().st_size
            print(f"テスト成功! ファイルサイズ: {file_size} bytes")
        else:
            print("テスト失敗: ファイルが作成されませんでした")
        
        print("\nExcel レポート生成テスト完了")
        
    except Exception as e:
        print(f"Excel テスト中にエラー: {e}")


if __name__ == "__main__":
    test_excel_generation()